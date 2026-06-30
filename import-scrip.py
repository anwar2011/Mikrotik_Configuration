import paramiko
import netmiko
from netmiko import ConnectHandler
import openpyxl

# Read command list
with open('commands_file') as f:
    commands_list = f.read().splitlines()

# Read device list
with open('devices_file') as f:
    devices_list = f.read().splitlines()

# Create Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MikroTik Results"

# Write headers
ws.cell(row=1, column=1, value="Device IP")
ws.cell(row=1, column=2, value="Command")
ws.cell(row=1, column=3, value="Output")

row = 2  # Start from row 2

for ip_address_of_device in devices_list:
    try:
        print('Connecting to device: ' + ip_address_of_device)

        device = {
            'device_type': 'mikrotik_routeros',
            'ip': ip_address_of_device,
            'username': 'bras',
            'password': 'OmlamP2024||'
        }

        net_connect = ConnectHandler(**device)

        # Run all commands from file
        for command in commands_list:
            output = net_connect.send_command(command)

            print(output)  # show on console
            print("Successful!")

            # Write to Excel
            ws.cell(row=row, column=1, value=ip_address_of_device)
            ws.cell(row=row, column=2, value=command)
            ws.cell(row=row, column=3, value=output)
            row += 1

        net_connect.disconnect()

    except Exception as e:
        print("failed")
        with open("failed_device.txt", "a+") as fail_dev:
            fail_dev.write(ip_address_of_device + '\n')
        continue

# Save Excel file
wb.save("mikrotik_output.xlsx")
print("📂 All results saved to mikrotik_output.xlsx")
